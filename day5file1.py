import openpyxl as xl
wb = xl.load_workbook('Employee.xlsx')
print('Type of wb:', type(wb))
print('Active worksheet object :', wb.active)
print('Worksheet names :', wb.sheetnames)
print('Worksheet objetcs :', wb.worksheets)
ws = wb['Emp Personal Data']
print('Type of ws:', type(ws))

#Fetching cell
print(ws.cell(1,1).value)
cell = ws['A2']
print(cell.value)

# slicing worksheet
cells = ws['A1':'C4']
for row in cells:
    for col in row:
        print(col.value, end=" ")
    print()

# max columns and rows in use
print(ws.max_column, ws.max_row)

# Accessing data row wise and column wise
rows = tuple(ws.rows)
cols = tuple(ws.columns)
print("Accessing rows.")
# for row in rows:
#     for col in row:
#         print(col.value, end=" ")
#     print()

# Fetching values for matching rows
emp_id_to_match = 1196035
flag = 0
for row in range(2, ws.max_row+1):
    if ws.cell(row,1).value == emp_id_to_match:  
        cells = ws[row]
        flag = 1
        break
if flag:
    for cell in cells:
        print(cell.value, end=' ')
else:
    print('Employee detail does not exist')

# #Adding as a list
# ws.append( [4, 'Payal', 10000] )
# #Adding as a tuple
# ws.append( (5, 'David', 10000) )
# #Saving the changes
# wb.save('Employee1.xlsx')

#Updating value
#Accessing the cell directly
ws['B4'] = 'James'
#Accessing through cell traversal
emp_id_to_update = 2
    
for row in range(2, ws.max_row+1):
    if ws.cell(row,1).value == emp_id_to_update:  
        ws.cell(row, 3).value = 50000
        break
else:
    print('Employee not found')
wb.save('Employees1.xlsx')